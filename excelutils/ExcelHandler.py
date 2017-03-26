import openpyxl
# 获取一个工作簿
def get_workbook(filename, read_only=False, keep_vba=False, data_only=False, guess_type=False, keep_link=True):
    wok = openpyxl.load_workbook(filename,read_only, keep_vba, data_only, guess_type, keep_link)
    return wok

# 获取工作簿里面所有的工作表名
def get_worksheet_namse_from_workbook(workbook):
    return workbook.get_sheet_names()

# 获取工作表
def get_worksheet(workbook, sheetname):
    return workbook.get_sheet_by_name(sheetname)

# 获得工作表行列数
def get_sheet_row_col_num(sheet):
    return (sheet.max_row,sheet.max_column)

# 获得指定单元格的值
def get_cell(sheet, location):
    return sheet[location].value

# 获得指定单元格的值
def get_cell(sheet, row, col):
    return sheet.cell(row, col).value

# 获得指定范围单元格的值
def get_cells(sheet, location_head, location_tail):
    return sheet[location_head:location_tail]

# 创建一个工作簿
def create_work(workbook_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'sheet1'
    workbook.save(workbook_name)

    return workbook

# 创建一个工作表
def create_sheet(workbook, sheet_name):
    workbook.create_sheet(title=sheet_name)

# 保存一个工作簿
def save(workbook, workbook_name):
    workbook.save(workbook_name)

# 以元组的形式获得表中所有行
def get_sheet_all_rows(sheet):
    return  list(sheet.rows)

def row_tostring(row):
    result = ""
    for cell in row:
        result = result +"."+cell.value
    return  result